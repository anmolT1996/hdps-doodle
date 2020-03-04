from django.shortcuts import render
from django.http import HttpResponseRedirect,HttpResponse
from django.urls import reverse_lazy,reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic.edit import CreateView,UpdateView,DeleteView
from django.views.generic.detail import DetailView,SingleObjectMixin
from django.views.generic.list import ListView
from hdps import forms
import openpyxl
from .models import Profile
from .forms import ProfileModelForm,SignUpForm,UserReportForm
from .models import UserReport
from hdps.apps import HdpsConfig
from sklearn.preprocessing import StandardScaler as ss
import numpy as np
# Create your views here.
class SignUp(CreateView):
    form_class = forms.SignUpForm
    success_url = reverse_lazy('hdps:login')
    template_name = 'signup.html'

class ProfileForm(LoginRequiredMixin,CreateView):
    form_class = forms.ProfileModelForm
    template_name = 'profile_form.html'

    def form_valid(self, form):
        fields = ['age','birth_date','phone_no','address','profile_pic']
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('hdps:profile-detail',kwargs={'pk':self.object.user.id})


class ProfileDetailView(DetailView):
    model = Profile
    template_name = 'profile_detail.html'

    def get_object(self):
        try:
            object = super().get_object()
        except:
            object = None
            return object
        return object



class ProfileUpdateView(UpdateView):
    model = Profile
    form_class = ProfileModelForm
    template_name = 'profile_form.html'

    def get_object(self):
        id_ = self.request.user.id
        return super().get_object()

    def form_valid(self, form):
        fields = ['age','birth_date','phone_no','address','profile_pic']
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('hdps:profile-detail',kwargs={'pk':self.object.user.id})

#############################################################################################

def userReportView(request):
    if request.method == 'POST':
        form = UserReportForm(request.POST)
        if form.is_valid():
            user = request.user
            age = form.cleaned_data['age']
            sex = form.cleaned_data['sex']
            cp = form.cleaned_data['cp']
            chol = form.cleaned_data['chol']
            fbs = form.cleaned_data['fbs']
            ex_in_angina = form.cleaned_data['ex_in_angina']
            st_depression_in_exercise = form.cleaned_data['st_depression_in_exercise']
            peak_st_segment = form.cleaned_data['peak_st_segment']
            vessels_by_flourosopy = form.cleaned_data['vessels_by_flourosopy']
            thal = form.cleaned_data['thal']

            excel_file = request.FILES["excel_file"]
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            restbps = worksheet.cell(row=2,column=1).value
            ecg =  worksheet.cell(row=2,column=2).value
            heart_rate =  worksheet.cell(row=2,column=3).value
            u = UserReport(user=user,age=age,sex=sex,cp=cp,restbps=restbps,chol=chol,fbs=fbs,ecg=ecg,heart_rate=heart_rate,ex_in_angina=ex_in_angina,st_depression_in_exercise=st_depression_in_exercise,peak_st_segment=peak_st_segment,vessels_by_flourosopy=vessels_by_flourosopy,thal=thal)
            u.save()
            return HttpResponseRedirect(reverse('hdps:heart_prediction'))
    else:
        form =UserReportForm()
    return render(request,'report_form.html',{'form':form})
############################################################################################
def home(request):
    return render(request,'home.html')
##########################################################################################
def predict(request):
    data = UserReport.objects.filter(user=request.user).latest('date_time')
    print(data.age)
    reportData = np.array([data.age,data.sex,data.cp,data.restbps,data.chol,data.fbs,data.ecg,data.heart_rate,data.ex_in_angina,data.st_depression_in_exercise,data.peak_st_segment,data.vessels_by_flourosopy,data.thal])
    reportData = reportData.reshape(1,-1)
    sc = ss()
    trained_data = sc.fit_transform(HdpsConfig.trained_data)
    reportData = sc.transform(reportData)
    result = HdpsConfig.classifier.predict(reportData)
    if data.sex == 1:
        sex = 'MALE'
    else:
        sex= 'FEMALE'
    if data.ecg == 0:
        ecg_val = 'Normal'
    elif data.ecg == 1:
        ecg_val = 'ST-T Wave Abnormality'
    else:
        ecg_val = 'Left Ventricular Hyperthrophy'

    try:
        profile = Profile.objects.get(user = request.user)
    except:
        profile = None
    return render(request,'prediction_report.html',{'report_data':data,'result':result,'sex':sex,'profile':profile,'ecg_val':ecg_val})
